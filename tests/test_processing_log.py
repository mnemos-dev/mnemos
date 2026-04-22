import pytest
from pathlib import Path


def test_upsert_creates_file_with_headers(tmp_path):
    from mnemos.processing_log import HEADERS, upsert_row

    vault = tmp_path
    (vault / "Mnemos").mkdir()

    upsert_row(vault, source_type="jsonl", path=tmp_path / "a.jsonl",
               refine_outcome="PENDING")

    import openpyxl
    wb = openpyxl.load_workbook(vault / "Mnemos" / "_processing.xlsx")
    ws = wb.active
    assert [c.value for c in ws[1]] == HEADERS
    # Row 2 has path + source_type + refine_outcome; other cols empty / "-"
    row = {HEADERS[i]: ws.cell(row=2, column=i+1).value for i in range(len(HEADERS))}
    assert row["source_type"] == "jsonl"
    assert row["path"] == str(tmp_path / "a.jsonl")
    assert row["refine_outcome"] == "PENDING"


def test_upsert_updates_existing_row(tmp_path):
    from mnemos.processing_log import upsert_row, read_rows
    (tmp_path / "Mnemos").mkdir()

    jsonl = tmp_path / "a.jsonl"
    upsert_row(tmp_path, source_type="jsonl", path=jsonl, refine_outcome="PENDING")
    upsert_row(tmp_path, source_type="jsonl", path=jsonl,
               refined_at="2026-04-22T10:00:00Z", refine_outcome="OK")

    rows = read_rows(tmp_path)
    assert len(rows) == 1
    assert rows[0]["refine_outcome"] == "OK"
    assert rows[0]["refined_at"] == "2026-04-22T10:00:00Z"
    assert rows[0]["source_type"] == "jsonl"


def test_upsert_rejects_unknown_field(tmp_path):
    from mnemos.processing_log import upsert_row
    (tmp_path / "Mnemos").mkdir()

    with pytest.raises(ValueError, match="Unknown"):
        upsert_row(tmp_path, source_type="jsonl", path=tmp_path / "a.jsonl",
                   not_a_column="x")


def test_upsert_two_distinct_paths_creates_two_rows(tmp_path):
    from mnemos.processing_log import upsert_row, read_rows
    (tmp_path / "Mnemos").mkdir()

    upsert_row(tmp_path, source_type="jsonl", path=tmp_path / "a.jsonl")
    upsert_row(tmp_path, source_type="md", path=tmp_path / "b.md")

    rows = read_rows(tmp_path)
    assert {r["path"] for r in rows} == {
        str(tmp_path / "a.jsonl"), str(tmp_path / "b.md"),
    }


def test_upsert_concurrent_safe(tmp_path):
    """Two threads writing different rows concurrently must both land without corruption."""
    import threading
    from mnemos.processing_log import upsert_row, read_rows
    (tmp_path / "Mnemos").mkdir()

    def writer(i: int):
        upsert_row(tmp_path, source_type="jsonl",
                   path=tmp_path / f"s{i}.jsonl", refine_outcome="OK")

    threads = [threading.Thread(target=writer, args=(i,)) for i in range(8)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    rows = read_rows(tmp_path)
    assert len(rows) == 8
    assert all(r["refine_outcome"] == "OK" for r in rows)
