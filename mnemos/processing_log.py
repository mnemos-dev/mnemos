"""Human-readable aggregate view of refine + skill-mine outcomes.

This module writes <vault>/Mnemos/_processing.xlsx — a flat Excel workbook
where each row corresponds to a single source file (JSONL transcript or
refined Session .md) and columns carry the latest refine/mine state. It is
NOT a canonical ledger; the canonical records live in the skill state dirs
(~/.claude/skills/mnemos-{refine-transcripts,mine-llm}/state/*.tsv). This
file is a convenience aggregate that the hook and catch-up both update after
each subprocess call. Obsidian hides it via the `_` prefix; Windows
Explorer opens it directly in Excel.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from filelock import FileLock, Timeout
import openpyxl

HEADERS = [
    "source_type", "path", "refined_at", "refine_outcome",
    "mined_at", "mined_outcome", "drawer_count", "tokens", "notes",
]

XLSX_NAME = "_processing.xlsx"
LOCK_NAME = ".mnemos-processing-xlsx.lock"


def _paths(vault: Path) -> tuple[Path, Path]:
    mnemos_dir = Path(vault) / "Mnemos"
    return mnemos_dir / XLSX_NAME, mnemos_dir / LOCK_NAME


def _load_or_create(xlsx_path: Path):
    if xlsx_path.exists():
        return openpyxl.load_workbook(xlsx_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "processing"
    ws.append(HEADERS)
    return wb


def _find_row(ws, path_str: str) -> int | None:
    """Return 1-based row number whose `path` cell matches, or None."""
    path_col = HEADERS.index("path") + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=path_col).value == path_str:
            return r
    return None


def upsert_row(vault: Path, *, source_type: str, path: Path, **fields: Any) -> None:
    """Insert or update the row keyed by `path`.

    `source_type` is set only on initial insert (later upserts don't touch it).
    `fields` may include any of HEADERS minus `source_type` and `path`.
    Unknown field names raise ValueError. `None` values are skipped (preserves
    existing cell content).
    """
    allowed = set(HEADERS) - {"source_type", "path"}
    unknown = set(fields) - allowed
    if unknown:
        raise ValueError(f"Unknown processing_log fields: {unknown}")

    xlsx_path, lock_path = _paths(vault)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    path_str = str(Path(path))

    lock = FileLock(str(lock_path), timeout=5)
    with lock:
        wb = _load_or_create(xlsx_path)
        ws = wb.active
        row_num = _find_row(ws, path_str)
        if row_num is None:
            row_values = [None] * len(HEADERS)
            row_values[HEADERS.index("source_type")] = source_type
            row_values[HEADERS.index("path")] = path_str
            for k, v in fields.items():
                if v is None:
                    continue
                row_values[HEADERS.index(k)] = v
            ws.append(row_values)
        else:
            for k, v in fields.items():
                if v is None:
                    continue
                ws.cell(row=row_num, column=HEADERS.index(k) + 1).value = v
        wb.save(xlsx_path)


def read_rows(vault: Path) -> list[dict]:
    """Return every row as a list of dicts keyed by HEADERS. Empty if file missing."""
    xlsx_path, _ = _paths(vault)
    if not xlsx_path.exists():
        return []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row = {HEADERS[i]: ws.cell(row=r, column=i+1).value for i in range(len(HEADERS))}
        if row["path"] is None:
            continue
        out.append(row)
    return out
